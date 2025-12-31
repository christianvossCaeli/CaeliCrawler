"""Summary Export Service.

Handles export of custom summaries to PDF and Excel formats.
Uses WeasyPrint for PDF generation and openpyxl for Excel.
"""

import io
import re
import unicodedata
from datetime import datetime
from typing import Any
from uuid import UUID

import structlog
from jinja2 import BaseLoader, Environment


def sanitize_filename(filename: str, max_length: int = 100) -> str:
    """
    Sanitize a filename to prevent path traversal and injection attacks.

    Args:
        filename: The original filename (without extension)
        max_length: Maximum length of the sanitized filename

    Returns:
        A safe filename string
    """
    if not filename:
        return "export"

    # Normalize unicode characters
    filename = unicodedata.normalize("NFKD", filename)
    filename = filename.encode("ascii", "ignore").decode("ascii")

    # Remove path traversal attempts
    filename = filename.replace("..", "")
    filename = filename.replace("/", "_")
    filename = filename.replace("\\", "_")

    # Remove header injection characters
    filename = filename.replace("\n", "")
    filename = filename.replace("\r", "")
    filename = filename.replace("\x00", "")

    # Remove other potentially dangerous characters
    # Keep only alphanumeric, spaces, hyphens, underscores
    filename = re.sub(r'[^\w\s\-]', '', filename)

    # Replace multiple spaces/underscores with single
    filename = re.sub(r'[\s_]+', '_', filename)

    # Strip leading/trailing underscores
    filename = filename.strip('_')

    # Truncate to max length
    if len(filename) > max_length:
        filename = filename[:max_length].rstrip('_')

    return filename or "export"


from sqlalchemy import select  # noqa: E402
from sqlalchemy.ext.asyncio import AsyncSession  # noqa: E402
from sqlalchemy.orm import selectinload  # noqa: E402

from app.models import CustomSummary, SummaryExecution, SummaryWidget  # noqa: E402
from app.models.summary_execution import ExecutionStatus  # noqa: E402

logger = structlog.get_logger(__name__)

# PDF Template using HTML/CSS (rendered by WeasyPrint)
PDF_TEMPLATE = """
<!DOCTYPE html>
<html lang="de">
<head>
    <meta charset="UTF-8">
    <title>{{ summary.name }}</title>
    <style>
        @page {
            size: A4 landscape;
            margin: 1.5cm;
            @bottom-center {
                content: "Seite " counter(page) " von " counter(pages);
                font-size: 9px;
                color: #666;
            }
        }

        * {
            box-sizing: border-box;
        }

        body {
            font-family: 'DejaVu Sans', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            font-size: 9pt;
            line-height: 1.4;
            color: #333;
            margin: 0;
            padding: 0;
        }

        .header {
            border-bottom: 2px solid #113534;
            padding-bottom: 12px;
            margin-bottom: 15px;
        }

        .header h1 {
            color: #113534;
            font-size: 20pt;
            margin: 0 0 6px 0;
        }

        .header .subtitle {
            color: #666;
            font-size: 10pt;
        }

        .meta {
            background: #f5f5f5;
            padding: 8px 12px;
            border-radius: 4px;
            margin-bottom: 15px;
            font-size: 9pt;
            color: #666;
        }

        .meta span {
            margin-right: 15px;
        }

        .widget {
            margin-bottom: 20px;
            border: 1px solid #e0e0e0;
            border-radius: 4px;
            overflow: hidden;
            page-break-inside: avoid;
        }

        .widget-header {
            background: #113534;
            color: white;
            padding: 8px 12px;
        }

        .widget-header h2 {
            margin: 0;
            font-size: 12pt;
        }

        .widget-header .subtitle {
            font-size: 9pt;
            opacity: 0.8;
            margin-top: 2px;
        }

        .widget-content {
            padding: 12px;
        }

        table {
            width: 100%;
            border-collapse: collapse;
            font-size: 8pt;
        }

        th {
            background: #113534;
            color: white;
            text-align: left;
            padding: 6px 8px;
            font-weight: 600;
            white-space: nowrap;
        }

        td {
            padding: 5px 8px;
            border-bottom: 1px solid #eee;
            vertical-align: top;
            word-wrap: break-word;
            max-width: 200px;
        }

        tr:nth-child(even) {
            background: #f9f9f9;
        }

        tr:hover {
            background: #f0f0f0;
        }

        .stat-card {
            text-align: center;
            padding: 15px;
        }

        .stat-value {
            font-size: 32pt;
            font-weight: bold;
            color: #113534;
        }

        .stat-label {
            font-size: 11pt;
            color: #666;
            margin-top: 4px;
        }

        .text-widget {
            white-space: pre-wrap;
        }

        .footer {
            margin-top: 20px;
            padding-top: 10px;
            border-top: 1px solid #ddd;
            font-size: 8pt;
            color: #999;
            text-align: center;
        }

        .no-data {
            color: #999;
            font-style: italic;
            text-align: center;
            padding: 15px;
        }

        .chart-info {
            background: #e8f4f8;
            padding: 10px 15px;
            border-radius: 4px;
            margin-bottom: 10px;
            font-size: 9pt;
            color: #0d6efd;
        }

        .data-count {
            font-size: 8pt;
            color: #666;
            margin-bottom: 8px;
        }

        .truncated-note {
            text-align: center;
            font-style: italic;
            color: #666;
            padding: 8px;
            background: #f5f5f5;
        }
    </style>
</head>
<body>
    <div class="header">
        <h1>{{ summary.name }}</h1>
        {% if summary.description %}
        <div class="subtitle">{{ summary.description }}</div>
        {% endif %}
    </div>

    <div class="meta">
        <span><strong>Erstellt:</strong> {{ export_date }}</span>
        {% if last_executed %}
        <span><strong>Letzte Aktualisierung:</strong> {{ last_executed }}</span>
        {% endif %}
        <span><strong>Widgets:</strong> {{ widgets|length }}</span>
    </div>

    {% for widget in widgets %}
    <div class="widget">
        <div class="widget-header">
            <h2>{{ widget.title }}</h2>
            {% if widget.subtitle %}
            <div class="subtitle">{{ widget.subtitle }}</div>
            {% endif %}
        </div>
        <div class="widget-content">
            {% set widget_data = cached_data.get('widget_' ~ widget.id) %}

            {% if widget.widget_type == 'stat_card' %}
                {% if widget_data and widget_data.data %}
                <div class="stat-card">
                    <div class="stat-value">{{ widget_data.data[0].value if widget_data.data else '-' }}</div>
                    <div class="stat-label">{{ widget.title }}</div>
                </div>
                {% else %}
                <div class="no-data">Keine Daten verfuegbar</div>
                {% endif %}

            {% elif widget.widget_type == 'text' %}
                <div class="text-widget">{{ widget_data.data[0].text if widget_data and widget_data.data else 'Kein Text' }}</div>

            {% elif widget.widget_type in ['bar_chart', 'line_chart', 'pie_chart'] %}
                <div class="chart-info">
                    Diagramm-Visualisierung ({{ widget.widget_type | replace('_', ' ') | title }}) - Daten als Tabelle exportiert
                </div>
                {% if widget_data and widget_data.data %}
                <div class="data-count">{{ widget_data.total }} Datensaetze</div>
                {{ render_data_table(widget_data.data, widget_data.total, 100, widget.visualization_config.get('columns') if widget.visualization_config else none) }}
                {% else %}
                <div class="no-data">Keine Daten verfuegbar</div>
                {% endif %}

            {% elif widget.widget_type == 'table' %}
                {% if widget_data and widget_data.data %}
                <div class="data-count">{{ widget_data.total }} Datensaetze</div>
                {{ render_data_table(widget_data.data, widget_data.total, 200, widget.visualization_config.get('columns') if widget.visualization_config else none) }}
                {% else %}
                <div class="no-data">Keine Daten verfuegbar</div>
                {% endif %}

            {% elif widget.widget_type == 'comparison' %}
                {% if widget_data and widget_data.data %}
                <div class="data-count">{{ widget_data.data | length }} Eintraege im Vergleich</div>
                {{ render_comparison_table(widget_data.data, widget.visualization_config.get('columns') if widget.visualization_config else none) }}
                {% else %}
                <div class="no-data">Keine Vergleichsdaten verfuegbar</div>
                {% endif %}

            {% elif widget.widget_type == 'map' %}
                {% if widget_data and widget_data.data %}
                {{ render_map_image(widget_data.data, widget_data.total, map_images.get('widget_' ~ widget.id|string)) }}
                {% else %}
                <div class="no-data">Keine Standortdaten verfuegbar</div>
                {% endif %}

            {% else %}
                {% if widget_data and widget_data.data %}
                <div class="data-count">{{ widget_data.total }} Datensaetze</div>
                {{ render_data_table(widget_data.data, widget_data.total, 100, widget.visualization_config.get('columns') if widget.visualization_config else none) }}
                {% else %}
                <div class="no-data">Keine Daten verfuegbar</div>
                {% endif %}
            {% endif %}
        </div>
    </div>
    {% endfor %}

    <div class="footer">
        Generiert von CaeliCrawler am {{ export_date }}
    </div>
</body>
</html>
"""


# Excluded keys from data table display (internal/technical fields)
EXCLUDED_DATA_KEYS = {'entity_id', 'coords_from_parent'}

def _format_key_label(key: str) -> str:
    """
    Format a key name into a human-readable label.
    Converts snake_case and kebab-case to Title Case.

    Examples:
        'admin_level_1' -> 'Admin Level 1'
        'power_mw' -> 'Power Mw'
        'wind-speed-ms' -> 'Wind Speed Ms'
    """
    if not key:
        return key
    # Replace underscores and hyphens with spaces
    label = key.replace('_', ' ').replace('-', ' ')
    # Title case
    return label.title()


def _format_value(value: Any) -> str:
    """Format a value for display in PDF."""
    if value is None:
        return '-'
    if isinstance(value, bool):
        return 'Ja' if value else 'Nein'
    if isinstance(value, float):
        # Format floats with 2 decimal places, remove trailing zeros
        if value == int(value):
            return str(int(value))
        return f'{value:.2f}'.rstrip('0').rstrip('.')
    if isinstance(value, dict):
        # Handle facet values
        if 'value' in value:
            return _format_value(value['value'])
        return str(value)
    if isinstance(value, list):
        return ', '.join(str(v) for v in value[:5])
    return str(value)


def _get_all_keys(data: list[dict]) -> list[str]:
    """Extract all unique keys from data, excluding internal keys."""
    if not data:
        return []

    keys = set()
    for row in data[:50]:  # Sample first 50 rows for performance
        for key in row:
            if key not in EXCLUDED_DATA_KEYS and key != 'facets':
                keys.add(key)

    # Sort with name first, then alphabetically
    sorted_keys = []
    if 'name' in keys:
        sorted_keys.append('name')
        keys.remove('name')
    sorted_keys.extend(sorted(keys))
    return sorted_keys


def _get_facet_keys(data: list[dict]) -> list[str]:
    """Extract all facet keys from data."""
    if not data:
        return []

    facet_keys = set()
    for row in data[:50]:
        facets = row.get('facets', {})
        if facets:
            facet_keys.update(facets.keys())
    return sorted(facet_keys)


def render_data_table(
    data: list[dict],
    total: int,
    max_rows: int = 100,
    columns_config: list[dict] | None = None,
) -> str:
    """
    Render a data table with configured or auto-detected columns.

    Args:
        data: List of data rows
        total: Total number of records
        max_rows: Maximum rows to display
        columns_config: Optional list of column configs with 'key' and 'label'
    """
    if not data:
        return '<div class="no-data">Keine Daten</div>'

    # Determine columns from config or auto-detect
    if columns_config and len(columns_config) > 0:
        # Use configured columns
        column_keys = [col.get("key", "") for col in columns_config]
        column_labels = [col.get("label") or _format_key_label(col.get("key", "")) for col in columns_config]
    else:
        # Auto-detect from data
        keys = _get_all_keys(data)
        facet_keys = _get_facet_keys(data)
        column_keys = keys + [f"facets.{fk}.value" for fk in facet_keys]
        column_labels = [_format_key_label(k) for k in keys] + [_format_key_label(fk) for fk in facet_keys]

    # Build header
    html = ['<table><thead><tr>']
    for label in column_labels:
        html.append(f'<th>{label}</th>')
    html.append('</tr></thead><tbody>')

    # Build rows
    display_data = data[:max_rows]
    for row in display_data:
        html.append('<tr>')
        for key in column_keys:
            value = _get_nested_value(row, key)
            html.append(f'<td>{_format_value(value)}</td>')
        html.append('</tr>')

    html.append('</tbody></table>')

    # Add truncation note if needed
    if total > max_rows:
        html.append(f'<div class="truncated-note">... und {total - max_rows} weitere Eintraege (Export auf {max_rows} begrenzt)</div>')

    return ''.join(html)


def _get_nested_value(data: dict, key: str) -> Any:
    """Get a value using dot notation for nested access."""
    if not key or not data:
        return None

    if "." in key:
        parts = key.split(".")
        value = data
        for part in parts:
            if isinstance(value, dict):
                value = value.get(part)
            else:
                return None
        return value

    return data.get(key)


def render_map_image(data: list[dict], total: int, pregenerated_image: str | None = None) -> str:
    """Render a map visualization with a pre-generated screenshot.

    Styled to match the website's MapVisualization component with Caeli brand colors.

    Args:
        data: List of data rows with latitude/longitude
        total: Total number of data points
        pregenerated_image: Base64-encoded PNG image of the map (pre-generated asynchronously)
    """
    if not data:
        return '<div class="no-data">Keine Standortdaten</div>'

    # Count valid coordinates
    point_count = 0
    for row in data:
        lat = row.get('latitude')
        lng = row.get('longitude')
        if lat is not None and lng is not None:
            try:
                float(lat)
                float(lng)
                point_count += 1
            except (ValueError, TypeError):
                continue

    if point_count == 0:
        return '<div class="no-data">Keine Koordinaten verfuegbar</div>'

    # Build centered HTML output - matching website styling
    html_parts = []

    # Centered container with fixed width matching map
    html_parts.append('<div style="page-break-inside: avoid; display: flex; flex-direction: column; align-items: center;">')

    if pregenerated_image:
        # Map image - exactly 700px wide, centered
        html_parts.append(
            f'<img src="data:image/png;base64,{pregenerated_image}" '
            f'style="width: 700px; height: auto; border-radius: 8px; display: block;" />'
        )

        # Info card overlay simulation (top-left style info)
        html_parts.append(
            '<div style="width: 700px; margin-top: 8px; display: flex; justify-content: space-between; align-items: flex-start;">'
        )

        # Left: Feature count (like website overlay)
        html_parts.append(
            f'<div style="background: white; padding: 6px 10px; border-radius: 4px; box-shadow: 0 1px 3px rgba(0,0,0,0.12); font-size: 9pt;">'
            f'<strong>{point_count}</strong> Standorte auf der Karte'
            f'</div>'
        )

        # Right: Cluster legend (matching website exactly)
        html_parts.append(
            '<div style="background: white; padding: 6px 10px; border-radius: 4px; box-shadow: 0 1px 3px rgba(0,0,0,0.12);">'
            '<div style="font-size: 8pt; font-weight: 500; margin-bottom: 4px; color: #666;">Cluster-Legende</div>'
            '<div style="display: flex; gap: 12px; font-size: 8pt;">'
            # Small cluster (green)
            '<span style="display: inline-flex; align-items: center; gap: 4px;">'
            '<span style="width: 12px; height: 12px; background: #2E7D32; border-radius: 50%; border: 2px solid white; box-shadow: 0 0 0 1px #2E7D32;"></span>'
            '&lt; 10</span>'
            # Medium cluster (teal/Caeli)
            '<span style="display: inline-flex; align-items: center; gap: 4px;">'
            '<span style="width: 12px; height: 12px; background: #113634; border-radius: 50%; border: 2px solid white; box-shadow: 0 0 0 1px #113634;"></span>'
            '10-50</span>'
            # Large cluster (indigo)
            '<span style="display: inline-flex; align-items: center; gap: 4px;">'
            '<span style="width: 12px; height: 12px; background: #5c6bc0; border-radius: 50%; border: 2px solid white; box-shadow: 0 0 0 1px #5c6bc0;"></span>'
            '&gt; 50</span>'
            '</div>'
            '</div>'
        )

        html_parts.append('</div>')  # End flex container
    else:
        html_parts.append(
            '<div style="width: 700px; padding: 40px 20px; background: #f5f5f5; border-radius: 8px; text-align: center;">'
            f'<p style="color: #666; margin: 0;">Kartenansicht nicht verfuegbar - {point_count} Standorte</p>'
            '</div>'
        )

    html_parts.append('</div>')  # End centered container

    return ''.join(html_parts)


async def _generate_map_screenshot_async(points: list[dict], status_config: dict, default_color: str) -> str:
    """Generate a map screenshot using Playwright async API and Leaflet.

    Uses CartoDB Positron tiles and Caeli brand colors to match the website styling.
    """
    import base64
    import json
    import os
    import tempfile

    from playwright.async_api import async_playwright

    # Calculate bounds with padding
    lats = [p['lat'] for p in points]
    lngs = [p['lng'] for p in points]
    min_lat, max_lat = min(lats), max(lats)
    min_lng, max_lng = min(lngs), max(lngs)

    # Add some padding to bounds
    lat_padding = (max_lat - min_lat) * 0.1 or 0.01
    lng_padding = (max_lng - min_lng) * 0.1 or 0.01
    min_lat -= lat_padding
    max_lat += lat_padding
    min_lng -= lng_padding
    max_lng += lng_padding

    # Prepare markers JSON - use Caeli green for all markers (like website)
    caeli_green = '#2E7D32'
    markers_data = []
    for p in points[:500]:  # Limit markers
        markers_data.append({
            'lat': p['lat'],
            'lng': p['lng'],
            'name': p['name'][:50] if p['name'] else '',
        })

    # Create HTML with Leaflet map using CartoDB Positron tiles (like website)
    html_content = f'''<!DOCTYPE html>
<html>
<head>
    <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css" />
    <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
    <style>
        body {{ margin: 0; padding: 0; }}
        #map {{ width: 700px; height: 400px; }}
    </style>
</head>
<body>
    <div id="map"></div>
    <script>
        var map = L.map('map', {{ zoomControl: false, attributionControl: false }});

        // Use CartoDB Positron tiles (same as website light mode)
        L.tileLayer('https://{{s}}.basemaps.cartocdn.com/light_all/{{z}}/{{x}}/{{y}}{{r}}.png', {{
            maxZoom: 19,
            subdomains: 'abcd'
        }}).addTo(map);

        var bounds = [[{min_lat}, {min_lng}], [{max_lat}, {max_lng}]];
        map.fitBounds(bounds, {{ padding: [30, 30] }});

        var markers = {json.dumps(markers_data)};

        // Caeli brand colors
        var markerColor = '{caeli_green}';

        markers.forEach(function(m) {{
            L.circleMarker([m.lat, m.lng], {{
                radius: 8,
                fillColor: markerColor,
                color: '#ffffff',
                weight: 2,
                opacity: 1,
                fillOpacity: 0.9
            }}).addTo(map);
        }});

        // Signal that map is ready after tiles loaded
        setTimeout(function() {{
            window.mapReady = true;
        }}, 2500);
    </script>
</body>
</html>'''

    # Write HTML to temp file and take screenshot
    with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False) as f:
        f.write(html_content)
        html_path = f.name

    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            page = await browser.new_page(viewport={'width': 700, 'height': 400})
            await page.goto(f'file://{html_path}')

            # Wait for map tiles to load
            await page.wait_for_function('window.mapReady === true', timeout=15000)
            await page.wait_for_timeout(1000)  # Extra time for tiles

            # Take screenshot
            screenshot_bytes = await page.screenshot(type='png')
            await browser.close()

        return base64.b64encode(screenshot_bytes).decode('utf-8')
    finally:
        os.unlink(html_path)


def render_map_table(data: list[dict], total: int) -> str:
    """Render a map data table with coordinates and all attributes."""
    if not data:
        return '<div class="no-data">Keine Standortdaten</div>'

    # Get all keys, ensure lat/lng are included
    keys = _get_all_keys(data)
    facet_keys = _get_facet_keys(data)

    # Ensure coordinate columns are present
    if 'latitude' not in keys:
        keys.append('latitude')
    if 'longitude' not in keys:
        keys.append('longitude')

    # Build header
    html = ['<table><thead><tr>']
    for key in keys:
        label = _format_key_label(key)
        html.append(f'<th>{label}</th>')
    for fk in facet_keys:
        html.append(f'<th>{_format_key_label(fk)}</th>')
    html.append('</tr></thead><tbody>')

    # Build rows - show all for map data
    for row in data:
        html.append('<tr>')
        for key in keys:
            value = row.get(key)
            html.append(f'<td>{_format_value(value)}</td>')
        for fk in facet_keys:
            facets = row.get('facets', {})
            facet_val = facets.get(fk, {})
            html.append(f'<td>{_format_value(facet_val)}</td>')
        html.append('</tr>')

    html.append('</tbody></table>')

    if total > len(data):
        html.append(f'<div class="truncated-note">Hinweis: {total} Gesamteintraege, {len(data)} exportiert</div>')

    return ''.join(html)


def render_comparison_table(
    data: list[dict],
    columns_config: list[dict] | None = None,
) -> str:
    """
    Render a comparison table with entities as columns.

    Args:
        data: List of entity data to compare
        columns_config: Optional list of column configs with 'key' and 'label' (defines rows)
    """
    if not data:
        return '<div class="no-data">Keine Vergleichsdaten</div>'

    # Determine row keys from config or auto-detect
    if columns_config and len(columns_config) > 0:
        # Use configured columns as rows (excluding name which is used as header)
        row_configs = [c for c in columns_config if c.get("key") != "name"]
    else:
        # Auto-detect: Get all attribute keys and facet keys
        attr_keys = _get_all_keys(data)
        if 'name' in attr_keys:
            attr_keys.remove('name')
        facet_keys = _get_facet_keys(data)

        row_configs = [{"key": k, "label": _format_key_label(k)} for k in attr_keys]
        row_configs += [{"key": f"facets.{fk}.value", "label": _format_key_label(fk)} for fk in facet_keys]

    # Build header with entity names
    html = ['<table><thead><tr><th>Merkmal</th>']
    for item in data[:10]:
        html.append(f'<th>{item.get("name", "?")}</th>')
    html.append('</tr></thead><tbody>')

    # Build rows for each configured attribute
    for row_config in row_configs:
        key = row_config.get("key", "")
        label = row_config.get("label") or _format_key_label(key)
        html.append(f'<tr><td><strong>{label}</strong></td>')
        for item in data[:10]:
            value = _get_nested_value(item, key)
            html.append(f'<td>{_format_value(value)}</td>')
        html.append('</tr>')

    html.append('</tbody></table>')
    return ''.join(html)


class SummaryExportService:
    """
    Handles export of custom summaries to various formats.

    Supports:
    - PDF export using WeasyPrint
    - Excel export using openpyxl
    """

    def __init__(self, session: AsyncSession):
        self.session = session
        self.jinja_env = Environment(loader=BaseLoader())  # noqa: S701
        # Register helper functions as globals
        self.jinja_env.globals['render_data_table'] = render_data_table
        self.jinja_env.globals['render_map_table'] = render_map_table
        self.jinja_env.globals['render_map_image'] = render_map_image
        self.jinja_env.globals['render_comparison_table'] = render_comparison_table

    async def export_to_pdf(
        self,
        summary_id: UUID,
        execution_id: UUID | None = None,
    ) -> bytes:
        """
        Export summary to PDF format.

        Args:
            summary_id: ID of the summary to export
            execution_id: Optional specific execution to export
                         (defaults to latest completed)

        Returns:
            PDF file as bytes

        Raises:
            ValueError: If summary not found
            ImportError: If WeasyPrint is not installed
        """
        logger.info(
            "pdf_export_started",
            summary_id=str(summary_id),
            execution_id=str(execution_id) if execution_id else None,
        )

        try:
            from weasyprint import HTML
        except ImportError:
            logger.error("WeasyPrint not installed")
            raise ImportError(
                "WeasyPrint is required for PDF export. "
                "Install it with: pip install weasyprint"
            ) from None

        # Load summary with widgets
        result = await self.session.execute(
            select(CustomSummary)
            .options(selectinload(CustomSummary.widgets))
            .where(CustomSummary.id == summary_id)
        )
        summary = result.scalar_one_or_none()

        if not summary:
            raise ValueError(f"Summary {summary_id} not found")

        # Get execution data
        execution = await self._get_execution(summary_id, execution_id)
        cached_data = execution.cached_data if execution else {}

        # Sort widgets by position
        widgets = sorted(
            summary.widgets,
            key=lambda w: (w.position_y or 0, w.position_x or 0)
        )

        # Pre-generate map images asynchronously
        map_images = await self._generate_map_images(widgets, cached_data)

        # Render template
        template = self.jinja_env.from_string(PDF_TEMPLATE)
        html_content = template.render(
            summary=summary,
            widgets=widgets,
            cached_data=cached_data,
            map_images=map_images,
            export_date=datetime.now().strftime("%d.%m.%Y %H:%M"),
            last_executed=execution.completed_at.strftime("%d.%m.%Y %H:%M") if execution and execution.completed_at else None,
        )

        # Generate PDF
        html = HTML(string=html_content)
        pdf_bytes = html.write_pdf()

        logger.info(
            "summary_exported_pdf",
            summary_id=str(summary_id),
            size_bytes=len(pdf_bytes),
        )

        return pdf_bytes

    async def export_to_excel(
        self,
        summary_id: UUID,
        execution_id: UUID | None = None,
    ) -> bytes:
        """
        Export summary to Excel format.

        Each widget becomes a separate sheet in the workbook.

        Args:
            summary_id: ID of the summary to export
            execution_id: Optional specific execution to export

        Returns:
            Excel file (.xlsx) as bytes

        Raises:
            ValueError: If summary not found
            ImportError: If openpyxl is not installed
        """
        logger.info(
            "excel_export_started",
            summary_id=str(summary_id),
            execution_id=str(execution_id) if execution_id else None,
        )

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl not installed")
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            ) from None

        # Load summary with widgets
        result = await self.session.execute(
            select(CustomSummary)
            .options(selectinload(CustomSummary.widgets))
            .where(CustomSummary.id == summary_id)
        )
        summary = result.scalar_one_or_none()

        if not summary:
            raise ValueError(f"Summary {summary_id} not found")

        # Get execution data
        execution = await self._get_execution(summary_id, execution_id)
        cached_data = execution.cached_data if execution else {}

        # Create workbook
        wb = Workbook()

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="113534", end_color="113534", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Sort widgets by position
        widgets = sorted(
            summary.widgets,
            key=lambda w: (w.position_y or 0, w.position_x or 0)
        )

        # Create overview sheet
        overview_sheet = wb.active
        overview_sheet.title = "Uebersicht"

        # Summary info section
        overview_data = [
            ["Zusammenfassung", summary.name],
            ["Beschreibung", summary.description or "-"],
            ["Exportiert am", datetime.now().strftime("%d.%m.%Y %H:%M")],
            ["Letzte Aktualisierung", execution.completed_at.strftime("%d.%m.%Y %H:%M") if execution and execution.completed_at else "-"],
            ["Anzahl Widgets", len(widgets)],
            ["Status", summary.status.value],
        ]

        for row_idx, (label, value) in enumerate(overview_data, start=1):
            overview_sheet.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            overview_sheet.cell(row=row_idx, column=2, value=value)

        # Widget summary table
        widget_start_row = len(overview_data) + 3
        overview_sheet.cell(row=widget_start_row, column=1, value="Widget-Uebersicht").font = Font(bold=True, size=12)

        # Widget table headers
        widget_headers = ["Nr.", "Titel", "Typ", "Datensaetze", "Status"]
        for col_idx, header in enumerate(widget_headers, start=1):
            cell = overview_sheet.cell(row=widget_start_row + 1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Widget rows
        for idx, widget in enumerate(widgets):
            widget_key = f"widget_{widget.id}"
            widget_data = cached_data.get(widget_key, {})
            total = widget_data.get("total", 0)
            has_error = "error" in widget_data

            row = widget_start_row + 2 + idx
            overview_sheet.cell(row=row, column=1, value=idx + 1).border = thin_border
            overview_sheet.cell(row=row, column=2, value=widget.title).border = thin_border
            overview_sheet.cell(row=row, column=3, value=widget.widget_type.replace("_", " ").title()).border = thin_border
            overview_sheet.cell(row=row, column=4, value=total).border = thin_border
            status_text = "Fehler" if has_error else ("OK" if total > 0 else "Keine Daten")
            overview_sheet.cell(row=row, column=5, value=status_text).border = thin_border

        overview_sheet.column_dimensions['A'].width = 25
        overview_sheet.column_dimensions['B'].width = 50
        overview_sheet.column_dimensions['C'].width = 20
        overview_sheet.column_dimensions['D'].width = 15
        overview_sheet.column_dimensions['E'].width = 15

        # Create sheet for each widget
        for idx, widget in enumerate(widgets):
            widget_key = f"widget_{widget.id}"
            widget_data = cached_data.get(widget_key, {})
            data = widget_data.get("data", [])
            total = widget_data.get("total", len(data) if data else 0)

            # Create sheet with sanitized name
            sheet_name = self._sanitize_sheet_name(f"{idx + 1}. {widget.title}")
            ws = wb.create_sheet(title=sheet_name)

            # Widget title
            ws.cell(row=1, column=1, value=widget.title).font = Font(bold=True, size=14)
            if widget.subtitle:
                ws.cell(row=2, column=1, value=widget.subtitle).font = Font(italic=True, color="666666")

            # Add data count info
            ws.cell(row=3, column=1, value=f"Datensaetze: {total}").font = Font(italic=True, color="666666")

            start_row = 5

            if not data:
                ws.cell(row=start_row, column=1, value="Keine Daten verfuegbar")
                continue

            # Determine columns based on widget type
            if widget.widget_type == "stat_card":
                ws.cell(row=start_row, column=1, value="Wert")
                ws.cell(row=start_row, column=2, value=data[0].get("value", "-"))
                ws.cell(row=start_row, column=1).font = Font(bold=True)
                continue

            if widget.widget_type == "text":
                ws.cell(row=start_row, column=1, value=data[0].get("text", ""))
                continue

            # Table-like widgets - determine columns from config or auto-detect
            configured_columns = self._get_configured_columns(widget, data)

            # Build headers and keys from configuration
            headers = [col["label"] for col in configured_columns]
            header_keys = [col["key"] for col in configured_columns]

            # Write headers
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=start_row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Data rows - write configured columns
            for row_idx, row_data in enumerate(data, start=start_row + 1):
                for col_idx, key in enumerate(header_keys, start=1):
                    value = self._get_value_by_key(row_data, key)
                    cell = ws.cell(row=row_idx, column=col_idx, value=self._format_excel_value(value))
                    cell.border = thin_border

            # Auto-size columns based on content
            for col_idx, header in enumerate(headers, start=1):
                col_letter = get_column_letter(col_idx)
                max_width = len(str(header)) + 2
                for row_data in data[:20]:
                    key = header_keys[col_idx - 1]
                    value = str(self._get_value_by_key(row_data, key) or "")
                    max_width = max(max_width, min(len(value) + 2, 50))
                ws.column_dimensions[col_letter].width = max_width

            # Add note if data was truncated
            if total > len(data):
                note_row = start_row + len(data) + 2
                ws.cell(row=note_row, column=1, value=f"Hinweis: {total - len(data)} weitere Eintraege nicht exportiert").font = Font(italic=True, color="999999")

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        excel_bytes = output.getvalue()

        logger.info(
            "summary_exported_excel",
            summary_id=str(summary_id),
            size_bytes=len(excel_bytes),
            widgets=len(widgets),
        )

        return excel_bytes

    async def _get_execution(
        self,
        summary_id: UUID,
        execution_id: UUID | None = None,
    ) -> SummaryExecution | None:
        """Get execution data for export."""
        if execution_id:
            result = await self.session.execute(
                select(SummaryExecution).where(
                    SummaryExecution.id == execution_id,
                    SummaryExecution.summary_id == summary_id,
                )
            )
            return result.scalar_one_or_none()

        # Get latest completed execution
        result = await self.session.execute(
            select(SummaryExecution)
            .where(
                SummaryExecution.summary_id == summary_id,
                SummaryExecution.status == ExecutionStatus.COMPLETED,
            )
            .order_by(SummaryExecution.created_at.desc())
            .limit(1)
        )
        return result.scalar_one_or_none()

    def _sanitize_sheet_name(self, name: str) -> str:
        """Sanitize sheet name for Excel (max 31 chars, no special chars)."""
        # Remove invalid characters
        invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
        for char in invalid_chars:
            name = name.replace(char, '')

        # Truncate to 31 characters
        if len(name) > 31:
            name = name[:28] + "..."

        return name or "Sheet"

    def _format_excel_value(self, value: Any) -> Any:
        """Format a value for Excel export."""
        if value is None:
            return "-"
        if isinstance(value, bool):
            return "Ja" if value else "Nein"
        if isinstance(value, float):
            # Keep floats as numbers for Excel (allows formatting/calculations)
            if value == int(value):
                return int(value)
            return round(value, 4)
        if isinstance(value, dict):
            # Handle facet values or nested dicts
            if "value" in value:
                return self._format_excel_value(value["value"])
            # Serialize other dicts as string
            return str(value)
        if isinstance(value, list):
            # Join lists with comma
            return ", ".join(str(v) for v in value[:10])
        return value

    def _get_configured_columns(
        self,
        widget: SummaryWidget,
        data: list[dict[str, Any]],
    ) -> list[dict[str, str]]:
        """
        Get column configuration from widget or auto-detect from data.

        If widget has visualization_config.columns, use those (in order).
        Otherwise, auto-detect columns from data.

        Returns:
            List of dicts with 'key' and 'label' for each column
        """
        # Check for configured columns in visualization_config
        viz_config = widget.visualization_config or {}
        configured = viz_config.get("columns")

        if configured and isinstance(configured, list) and len(configured) > 0:
            # Use configured columns
            return [
                {
                    "key": col.get("key") or col.get("field", ""),
                    "label": col.get("label") or _format_key_label(col.get("key", "")),
                }
                for col in configured
                if col.get("key") or col.get("field")
            ]

        # Auto-detect from data
        columns = []

        # Get attribute keys
        all_keys = _get_all_keys(data)
        for key in all_keys:
            columns.append({
                "key": key,
                "label": _format_key_label(key),
            })

        # Get facet keys
        facet_keys = _get_facet_keys(data)
        for fk in facet_keys:
            columns.append({
                "key": f"facets.{fk}.value",
                "label": _format_key_label(fk),
            })

        return columns

    def _get_value_by_key(self, row_data: dict[str, Any], key: str) -> Any:
        """
        Get a value from row data using a key (supports dot notation for nested values).

        Examples:
            - 'name' -> row_data['name']
            - 'facets.points.value' -> row_data['facets']['points']['value']
        """
        if not key:
            return None

        # Handle dot notation for nested access
        if "." in key:
            parts = key.split(".")
            value = row_data
            for part in parts:
                if isinstance(value, dict):
                    value = value.get(part)
                else:
                    return None
            return value

        return row_data.get(key)

    async def _generate_map_images(
        self,
        widgets: list[SummaryWidget],
        cached_data: dict[str, Any],
    ) -> dict[str, str]:
        """
        Pre-generate map screenshots for all map widgets.

        Args:
            widgets: List of widgets to process
            cached_data: Cached widget data from execution

        Returns:
            Dict mapping widget_id to base64-encoded PNG image
        """
        map_images: dict[str, str] = {}

        # Status colors
        status_config = {
            'active': '#22c55e',
            'operational': '#22c55e',
            'operating': '#22c55e',
            'planned': '#3b82f6',
            'approved': '#a855f7',
            'construction': '#f97316',
            'under construction': '#f97316',
            'decommissioned': '#6b7280',
            'inactive': '#6b7280',
        }
        default_color = '#ef4444'

        for widget in widgets:
            if widget.widget_type != 'map':
                continue

            widget_key = f"widget_{widget.id}"
            widget_data = cached_data.get(widget_key, {})
            data = widget_data.get("data", [])

            if not data:
                continue

            # Extract coordinates
            points = []
            for row in data:
                lat = row.get('latitude')
                lng = row.get('longitude')
                if lat is not None and lng is not None:
                    try:
                        status = str(row.get('status', '')).lower().strip()
                        points.append({
                            'lat': float(lat),
                            'lng': float(lng),
                            'name': row.get('name', ''),
                            'status': status,
                        })
                    except (ValueError, TypeError):
                        continue

            if not points:
                continue

            # Generate screenshot asynchronously
            try:
                image_data = await _generate_map_screenshot_async(
                    points, status_config, default_color
                )
                map_images[widget_key] = image_data
                logger.info(
                    "map_screenshot_generated",
                    widget_id=str(widget.id),
                    points=len(points),
                )
            except Exception as e:
                logger.warning(
                    "map_screenshot_failed",
                    widget_id=str(widget.id),
                    error=str(e),
                )

        return map_images
